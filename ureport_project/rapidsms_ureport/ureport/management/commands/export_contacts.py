#!/usr/bin/python
# -*- coding: utf-8 -*-

from django.core.management.base import BaseCommand
import traceback
import os
from openpyxl.cell import get_column_letter
import xlwt
import datetime
from django.db import connection, transaction
from django.conf import settings

ezxf = xlwt.easyxf

from openpyxl.workbook import Workbook


class Command(BaseCommand):
    year_now = datetime.datetime.now().year

    sql = \
        """    SELECT
"rapidsms_contact"."id",
"rapidsms_contact"."language",
(SELECT
    "rapidsms_connection"."id"
 FROM
    "rapidsms_connection"
 WHERE
    "rapidsms_connection"."contact_id"="rapidsms_contact"."id") as id_connection,
(SELECT
    "rapidsms_connection"."identity"
 FROM
    "rapidsms_connection"
 WHERE
    "rapidsms_connection"."contact_id"="rapidsms_contact"."id") as identity,
"rapidsms_contact"."name",
"rapidsms_contact"."birthdate",
"rapidsms_contact"."created_on",


(SELECT
 DATE("rapidsms_httprouter_message"."date")
FROM
 "rapidsms_httprouter_message"
WHERE
 "rapidsms_httprouter_message"."direction" = 'I'
 and "rapidsms_httprouter_message"."application" = 'unregister'
 and  "rapidsms_httprouter_message"."connection_id" = (
    SELECT
       "rapidsms_connection"."id"
    FROM
       "rapidsms_connection"
    WHERE
       "rapidsms_connection"."contact_id" = "rapidsms_contact"."id"  LIMIT 1
 ) LIMIT 1
) as quit_date,

"locations_location"."name" as province,

(
 %d-EXTRACT('year'
FROM
 "rapidsms_contact"."birthdate")) as age,


 "rapidsms_contact"."gender",

 "rapidsms_contact"."health_facility" as facility,

  "rapidsms_contact"."colline_name" as colline,

(SELECT
    "locations_location"."name"
 FROM
    "locations_location"
 WHERE
    "locations_location"."id"="rapidsms_contact"."colline_id") as location,


 (array(SELECT
    "auth_group"."name"
 FROM
    "auth_group"
 INNER JOIN
    "rapidsms_contact_groups"
       ON (
          "auth_group"."id" = "rapidsms_contact_groups"."group_id"
       )
 WHERE
    "rapidsms_contact_groups"."contact_id" = "rapidsms_contact"."id" order by "auth_group"."id" ))[1] as group,

(SELECT
 COUNT(*) FROM
    "poll_response"
 WHERE
    "poll_response"."contact_id"="rapidsms_contact"."id") as responses,


    (SELECT DISTINCT
 COUNT(*) FROM
    "poll_poll_contacts"
 WHERE
    "poll_poll_contacts"."contact_id"="rapidsms_contact"."id" GROUP BY "poll_poll_contacts"."contact_id") as questions,



(SELECT DISTINCT count(*)

FROM "rapidsms_httprouter_message"

WHERE  "rapidsms_httprouter_message"."direction" ='I'  and

"rapidsms_httprouter_message"."connection_id" = (
    SELECT
       "rapidsms_connection"."id"
    FROM
       "rapidsms_connection"
    WHERE
       "rapidsms_connection"."contact_id" = "rapidsms_contact"."id"  LIMIT 1
 ) ) as incoming



FROM
 "rapidsms_contact"
LEFT JOIN
 "locations_location"
    ON "rapidsms_contact"."reporting_location_id" = "locations_location"."id";
         """ \
        % year_now

    def write_with_openpyxl(self, filename, headers, data):
        wb = Workbook()
        ws = wb.worksheets[0]
        ws.title = 'Ureporters'
        row_x = 0
        for row in data:
            row_x += 1
            for col_x, value in enumerate(headers):
                col = get_column_letter(col_x)
                ws.cell('%s%s' % (col, row_x)).value = value


    def chunks(l, n):
        """ Yield successive n-sized chunks from l.
        """
        for i in xrange(0, len(l), n):
            yield l[i:i + n]


    def write_xls(self, file_name, sheet_name, headings, data, data_xfs,
                  heading_xf=ezxf('font: bold on; align: wrap on, vert centre, horiz center')):
        data = self.chunks(data, 65000)
        book = xlwt.Workbook()
        sn = 0
        for dat in data:
            sn += 1
            sheet = book.add_sheet(sheet_name + str(sn))
            rowx = 0
            for colx, value in enumerate(headings):
                sheet.write(rowx, colx, value, heading_xf)
            sheet.set_panes_frozen(True) # frozen headings instead of split panes
            sheet.set_horz_split_pos(rowx + 1) # in general, freeze after last heading row
            sheet.set_remove_splits(True) # if user does unfreeze, don't leave a split there
            for row in dat:
                rowx += 1
                for colx, value in enumerate(row):
                    sheet.write(rowx, colx, value, data_xfs[colx])
        book.save(file_name)

    @transaction.atomic
    def handle(self, **options):
        try:

            from uganda_common.utils import ExcelResponse

            excel_file_path = \
                os.path.join(os.path.join(os.path.join(settings.STATIC_ROOT,'ureport'), 'spreadsheets'),
                             'ureporters.xlsx')
            export_data_list = []

            # messages=Message.objects.select_related(depth=1)
            # black_listed=Blacklist.objects.values_list('connection__contact__pk',flat=True)
            # print  black_listed
            # black_list_messages=messages.filter(connection__contact__in=black_listed)
            # opt_words=settings.OPT_OUT_WORDS

            if connection.connection is None:
                cursor = connection.cursor()
            cursor = connection.connection.cursor(name='contacts')
            cursor.execute(self.sql)
            row_0 = [
                (
                    'Id',
                    'Language',
                    'Id Connection',
                    'Number',
                    'Name',
                    'Birthdate',
                    'Join Date',
                    'Quit Date',
                    'Province',
                    'Age',
                    'Gender',
                    'Health Facility',
                    'Colline', #village
                    'Commune', #Subconty
                    'Group',
                    'Number Of Responses',
                    'Number Of Questions Asked',
                    'Number of Incoming',
                )
            ]

            rows = row_0 + cursor.fetchall()
            kinds = "int text text text text date date date text int text text text text text text text text".split()
            kind_to_xf_map = {
                'date': ezxf(num_format_str='yyyy-mm-dd'),
                'int': ezxf(num_format_str='#,##0'),
                'text': ezxf(),
            }

            data_xfs = [kind_to_xf_map[k] for k in kinds]
            ExcelResponse(rows, output_name=excel_file_path,
                          write_to_file=True)
            #print rows
            # self.write_xls(excel_file_path, 'ureporters', row_0, rows, data_xfs)
        except Exception, exc:
            print traceback.format_exc(exc)
